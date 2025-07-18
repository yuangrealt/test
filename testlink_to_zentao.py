import xml.etree.ElementTree as ET
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

def remove_html_tags(text):
    """移除HTML标签并保留换行符"""
    if text is None:
        return ""
    
    # 保留换行标记
    text = text.replace('<p>', '\n').replace('</p>', '\n')
    text = text.replace('<br>', '\n').replace('<br/>', '\n')
    text = text.replace('&nbsp;', ' ')
    
    # 移除其他HTML标签
    clean = re.compile('<.*?>')
    return re.sub(clean, '', text).strip()

def parse_testlink_xml(xml_file):
    """解析TestLink XML文件"""
    tree = ET.parse(xml_file)
    root = tree.getroot()
    
    testcases = []
    module_stack = []  # 用于保存当前模块路径
    
    for testsuite in root.findall('.//testsuite'):
        # 处理模块层级
        module_name = testsuite.get('name') or "未命名模块"
        parent = testsuite.getparent()
        
        # 计算当前模块层级
        level = 0
        while parent is not None and parent.tag == 'testsuite':
            level += 1
            parent = parent.getparent()
        
        # 更新模块栈
        if level < len(module_stack):
            module_stack = module_stack[:level]
        module_stack.append(module_name)
        full_module = "/".join(module_stack)
        
        # 提取测试用例
        for testcase in testsuite.findall('testcase'):
            case_data = {
                'module': full_module,
                'title': testcase.get('name'),
                'steps': [],
                'preconditions': remove_html_tags(testcase.findtext('preconditions')),
                'summary': remove_html_tags(testcase.findtext('summary')),
                'importance': testcase.findtext('importance', '2')
            }
            
            # 处理步骤
            steps = testcase.find('steps')
            if steps is not None:
                for step in steps.findall('step'):
                    step_num = step.findtext('step_number', '').strip()
                    actions = remove_html_tags(step.findtext('actions'))
                    expected = remove_html_tags(step.findtext('expectedresults'))
                    
                    case_data['steps'].append({
                        'step_number': step_num,
                        'actions': actions,
                        'expected': expected
                    })
            
            testcases.append(case_data)
    
    return testcases

def create_zentao_xlsx(testcases, output_file):
    """创建禅道格式的XLSX文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"
    
    # 设置标题行
    headers = [
        "所属模块", "用例标题", "前置条件", 
        "步骤", "预期", "优先级", "用例类型", 
        "适用阶段", "关键词"
    ]
    ws.append(headers)
    
    # 设置标题样式
    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold_font
    
    # 添加测试用例数据
    for case in testcases:
        # 组合步骤和预期
        steps_str = ""
        for i, step in enumerate(case['steps'], 1):
            steps_str += f"{i}. {step['actions']}\n"
        
        expected_str = ""
        for i, step in enumerate(case['steps'], 1):
            expected_str += f"{i}. {step['expected']}\n"
        
        # 优先级转换 (TestLink:1=高, 2=中, 3=低 → 禅道:1=低, 2=中, 3=高)
        priority_map = {'1': '3', '2': '2', '3': '1'}
        zentao_priority = priority_map.get(case['importance'], '2')
        
        # 添加行数据
        row = [
            case['module'],                      # 所属模块
            case['title'],                       # 用例标题
            case['preconditions'] or case['summary'],  # 前置条件
            steps_str.strip(),                   # 步骤
            expected_str.strip(),                # 预期
            zentao_priority,                     # 优先级
            "功能测试",                          # 用例类型
            "功能测试阶段",                      # 适用阶段
            ""                                   # 关键词
        ]
        ws.append(row)
    
    # 设置自动列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    
    # 设置文本换行
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap_alignment
    
    wb.save(output_file)
    print(f"成功生成禅道导入文件: {output_file}")

if __name__ == "__main__":
    input_xml = "testdemo.testproject-deep.xml"  # 输入的XML文件名
    output_xlsx = "zentao_import.xlsx"           # 输出的XLSX文件名
    
    testcases = parse_testlink_xml(input_xml)
    create_zentao_xlsx(testcases, output_xlsx)